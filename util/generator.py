import os, itertools, csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference


class LogFile:
    def __init__(self, file_path: str) -> None:
        self.file_path = file_path
        self.file_name = os.path.basename(file_path)
        self.scenario_name = self.file_name.split("_")[3]
        self.sub_scenario_name: str | None = None

        if "(" in self.file_name:
            first_index = self.file_name.index("(")
            last_index = self.file_name.index(")")
            self.sub_scenario_name = self.file_name[first_index + 1 : last_index]

    def __str__(self) -> str:
        return self.file_name


class Generator:
    def __init__(
        self,
        file_paths: list[str],
        # 분석 시점의 distanceTravelled 값 (meter)
        starting_point: str,
        # 분석 종점의 distanceTravelled 값 (meter)
        ending_point: str,
        # 분석 시점의 station 값 (kilometer)
        starting_station: str,
        # 분석 구간 빈도 (meter)
        frequency: str,
    ) -> None:
        self.starting_point = int(starting_point)
        self.ending_point = int(ending_point)
        self.starting_station = float(starting_station)
        self.frequency = int(frequency)
        self.frequency_in_kilometer = int(frequency) / 1000
        self.selected_columns = ["speedInKmPerHour", "offsetFromLaneCenter", "EEG"]

        log_files = [LogFile(file_path) for file_path in file_paths]
        self.grouped_log_files: list[list[LogFile]] = [log_files]

        if log_files[0].sub_scenario_name is not None:
            sorted_log_files = sorted(
                log_files, key=lambda file: file.sub_scenario_name
            )
            self.grouped_log_files = [
                list(g)
                for _, g in itertools.groupby(
                    sorted_log_files, key=lambda file: file.sub_scenario_name
                )
            ]

    def translate_selected_column_name(self, selected_column: str) -> str:
        translation: str = selected_column

        if selected_column == "offsetFromLaneCenter":
            translation = "차로편측"
        elif selected_column == "speedInKmPerHour":
            translation = "주행속도"
        elif selected_column == "EEG":
            translation = "뇌파"

        return translation

    def get_index_of_column(self, columns: list[str], target_column: str):
        target_column_index: int = -1

        for i, col_name in enumerate(columns):
            if col_name == target_column:
                target_column_index = i

        return target_column_index

    def represents_float(self, string: str) -> bool:
        try:
            float(string)
            return True
        except ValueError:
            return False

    def generate_workbook(self) -> Workbook:

        wb = Workbook()
        chart_sheet = wb.active
        chart_sheet.title = "그래프"

        chart_out_row = 0

        for group in self.grouped_log_files:
            chart_out_col = 0

            for selected_column in self.selected_columns:

                ws = wb.create_sheet()

                ws.title = self.translate_selected_column_name(selected_column)

                if group[0].sub_scenario_name is not None:
                    ws.title = f"{ws.title}_{group[0].sub_scenario_name}"

                # insert first two columns (STA, distanceTravelled)
                ws.append(["STA", "distanceTravelled"])
                row_num = 2
                for dt in range(
                    self.starting_point, self.ending_point + 1, self.frequency
                ):
                    ws.append(
                        [
                            self.starting_station * 1000 + dt - self.starting_point,
                            dt,
                        ]
                    )
                    ws[f"A{row_num}"].number_format = '0"+"000'
                    row_num += 1

                for log_file in group:
                    OUT_COL = ws.max_column + 1
                    out_row = 2

                    with open(log_file.file_path, encoding="utf8") as csv_file:
                        heading = csv_file.__next__()
                        if selected_column not in heading:
                            break

                        selected_column_index = self.get_index_of_column(
                            heading.split(","), selected_column
                        )
                        distanceTravelled_index = self.get_index_of_column(
                            heading.split(","), "distanceTravelled"
                        )

                        ws.cell(column=OUT_COL, row=1, value=OUT_COL - 2)

                        for dt in range(
                            self.starting_point, self.ending_point + 1, self.frequency
                        ):
                            closest_row = None

                            for row in csv.reader(csv_file):
                                if closest_row is None:
                                    closest_row = row
                                    continue

                                prev_difference = (
                                    float(closest_row[distanceTravelled_index]) - dt
                                )
                                difference = float(row[distanceTravelled_index]) - dt

                                if difference < -2:
                                    continue

                                if difference > 2:
                                    break

                                if abs(prev_difference) > abs(difference):
                                    closest_row = row

                            if closest_row is None:
                                raise Exception("분석 시점과 종점 값을 다시 확인해주세요")

                            closest_row_value: float | str

                            if self.represents_float(
                                closest_row[selected_column_index]
                            ):
                                closest_row_value = abs(
                                    float(closest_row[selected_column_index])
                                )
                            else:
                                closest_row_value = closest_row[selected_column_index]

                            ws.cell(
                                column=OUT_COL, row=out_row, value=closest_row_value
                            )
                            out_row += 1

                # 평균값 컬럼 생성
                AVERAGE_OUT_COL = ws.max_column + 1

                for row in range(1, ws.max_row + 1):
                    if row == 1:
                        ws.cell(row=row, column=AVERAGE_OUT_COL, value="평균")
                        continue

                    ws.cell(
                        row=row,
                        column=AVERAGE_OUT_COL,
                        value=f"=AVERAGE({'C' + str(row)}:{get_column_letter(AVERAGE_OUT_COL - 1) + str(row)})",
                    )

                # 차트 생성
                chart = LineChart()
                chart.title = ws.title
                chart.style = 2

                # y축 정보 입력
                chart.y_axis.title = self.translate_selected_column_name(
                    selected_column
                )
                data = Reference(
                    ws,
                    min_col=AVERAGE_OUT_COL,
                    min_row=2,
                    max_col=AVERAGE_OUT_COL,
                    max_row=ws.max_row,
                )
                chart.add_data(data, titles_from_data=True)

                # x축 정보 입력
                chart.x_axis.title = "STA. (km)"
                labels = Reference(
                    ws,
                    min_col=1,
                    min_row=2,
                    max_col=1,
                    max_row=ws.max_row,
                )
                chart.set_categories(labels) 

                chart.legend = None
                chart_sheet.add_chart(
                    chart,
                    f"{get_column_letter(2 + chart_out_col * 10)}{2 + chart_out_row * 15}",
                )

                chart_out_col += 1

            chart_out_row += 1
        return wb
