import os
import csv
from openpyxl import Workbook


class LogFile:
    def __init__(self, file_path) -> None:
        self.file_path = file_path
        self.file_name = os.path.basename(self.file_path)
        self.scenario_name = self.file_name.split("_")[3]

    def __str__(self) -> str:
        return self.file_name


file_paths = [
    "C:\\Users\\92060\\Documents\\python-projects\\Log_20211001131853_Unknown Road_시나리오1(1시간)_1_0_0_0.csv",
    "C:\\Users\\92060\\Documents\\python-projects\\Log_20211001132147_Unknown Road_시나리오1(1시간)_2_0_0_0.csv",
    "C:\\Users\\92060\\Documents\\python-projects\\Log_20211001132310_Unknown Road_시나리오1(1시간)_3_0_0_0.csv",
]

log_files = [LogFile(file_path) for file_path in file_paths]
selected_columns = ["speedInKmPerHour", "offsetFromLaneCenter"]

# 분석 시작점의 station 값 (km)
station = 14.2
station_cp = station

# 마지막 분석점의 distanceTravelled 값 (m)
ending_point = 1400

# 분석 시작점의 distanceTravelled 값 (m)
distance_travelled = 500
distance_travelled_cp = distance_travelled

# 분석 빈도 (m)
frequency_in_meter = 10
frequency_in_kilometer = frequency_in_meter / 1000

wb = Workbook()
ws = wb.active
ws.title = "주행 속도" if selected_columns[0] == "speedInKmPerHour" else "차로 편측"
ws.append(["STA", "distanceTravelled"])

row_num = 2

while distance_travelled_cp <= ending_point:
    ws.append([station_cp * 1000, distance_travelled_cp])
    ws[f"A{row_num}"].number_format = '0"+"000.00'
    row_num += 1
    station_cp += frequency_in_kilometer
    distance_travelled_cp += frequency_in_meter
else:
    distance_travelled_cp = distance_travelled
    station_cp = station


with open(log_files[0].file_path, encoding="utf8") as csv_file:

    heading = csv_file.__next__()

    output_col = ws.max_column + 1
    output_row = 2
    ws.cell(column=output_col, row=1, value=output_col - 2)

    for dt in range(distance_travelled, ending_point + 1, frequency_in_meter):
        closest_row = None

        for row in csv.reader(csv_file):
            if closest_row is None:
                closest_row = row
                continue

            prev_difference = float(closest_row[0]) - dt
            difference = float(row[0]) - dt

            if difference < -2:
                continue

            if difference > 2:
                break

            if abs(prev_difference) > abs(difference):
                closest_row = row

        ws.cell(column=output_col, row=output_row, value=closest_row[1])
        output_row += 1

wb.save("done.xlsx")
